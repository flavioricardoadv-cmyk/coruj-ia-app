"""
router.py â€” FastAPI Router do Motor de Aprendizado
===================================================
Cole este router no seu app FastAPI com:

    from motor_aprendizado import router as motor_router
    app.include_router(motor_router)

Prefixo padrÃ£o: /motor

Endpoints disponÃ­veis:
  POST /motor/upload                      â€” envia documento para aprendizado
  POST /motor/upload-com-resultado        â€” envia documento com resultado (aprendizado estratÃ©gico)
  GET  /motor/status                      â€” bateria de conhecimento por categoria
  GET  /motor/arvore                      â€” Ã¡rvore de conhecimento completa
  GET  /motor/perfil/{categoria}          â€” padrÃµes aprendidos de uma categoria
  GET  /motor/estrategia/{categoria}      â€” inteligÃªncia estratÃ©gica consolidada
  GET  /motor/fila                        â€” status da fila de processamento
  POST /motor/processar-agora             â€” forÃ§a processamento imediato
  GET  /motor/grafo/status                â€” estatÃ­sticas do Grafo de Conhecimento
  GET  /motor/grafo/{categoria}           â€” inteligÃªncia do grafo por categoria
  POST /motor/grafo/recalcular-pesos      â€” recalcula pesos temporais
  POST /motor/feedback/{doc_id}           â€” registra feedback sobre documento gerado
  GET  /motor/minutas                     â€” lista documentos gerados com mÃ©tricas
"""

import hashlib
import io
import re
from zipfile import ZipFile, BadZipFile
from datetime import datetime
from fastapi import APIRouter, UploadFile, File, Form, Depends, HTTPException
from sqlalchemy.orm import Session

from app.motor.models import get_db, MotorDocumento, MotorFila, MotorNivelConhecimento, MotorMinuta, MotorAprendizadoEstrategico
from app.motor import cerebro_service
from app.motor import kg_service
from app.motor.config import CATEGORIAS_VALIDAS

router = APIRouter(prefix="/motor", tags=["motor-aprendizado"])


# â”€â”€ ExtraÃ§Ã£o de texto (adapte para o seu projeto) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def _extrair_texto_arquivo(conteudo: bytes, nome_arquivo: str) -> str:
    """
    Extrai texto de um arquivo (PDF, DOCX, TXT, XLS, imagem).
    Por padrÃ£o tenta pdfplumber e python-docx.
    Substitua por seu prÃ³prio extrator se necessÃ¡rio.
    """
    ext = (nome_arquivo.split(".")[-1] if "." in nome_arquivo else "").lower()

    if ext == "rtf":
        return _extrair_texto_rtf(conteudo)

    if ext == "txt":
        for enc in ("utf-8", "latin-1", "cp1252"):
            try:
                return conteudo.decode(enc)
            except Exception:
                continue
        return conteudo.decode("utf-8", errors="replace")

    if ext == "pdf":
        try:
            import pdfplumber, io
            with pdfplumber.open(io.BytesIO(conteudo)) as pdf:
                return "\n".join(p.extract_text() or "" for p in pdf.pages)
        except Exception:
            pass

    if ext in ("docx",):
        try:
            from docx import Document
            import io
            doc = Document(io.BytesIO(conteudo))
            return "\n".join(p.text for p in doc.paragraphs)
        except Exception:
            pass

    if ext in ("xls", "xlsx"):
        text = _extrair_texto_planilha(conteudo, ext)
        if text:
            return text

    # Fallback: decodifica como texto
    return conteudo.decode("utf-8", errors="replace")


def _extrair_texto_planilha(conteudo: bytes, ext: str) -> str:
    """Converte planilhas exportadas pelo SAJ/Excel em linhas de texto treinaveis."""
    try:
        if ext == "xls":
            import xlrd
            book = xlrd.open_workbook(file_contents=conteudo)
            rows: list[str] = []
            for sheet in book.sheets():
                rows.append(f"Aba: {sheet.name}")
                for row_index in range(sheet.nrows):
                    values = []
                    for col_index in range(sheet.ncols):
                        value = sheet.cell_value(row_index, col_index)
                        if isinstance(value, float) and value.is_integer():
                            value = int(value)
                        values.append(str(value).strip())
                    line = " | ".join(value for value in values if value)
                    if line:
                        rows.append(line)
            return "\n".join(rows)

        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
        rows = []
        for sheet in workbook.worksheets:
            rows.append(f"Aba: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                line = " | ".join(str(value).strip() for value in row if value is not None and str(value).strip())
                if line:
                    rows.append(line)
        return "\n".join(rows)
    except Exception:
        return ""


def _extrair_texto_rtf(conteudo: bytes) -> str:
    """Extrator leve de RTF, suficiente para arquivos exportados por sistemas judiciais."""
    raw = ""
    for enc in ("utf-8", "cp1252", "latin-1"):
        try:
            raw = conteudo.decode(enc)
            break
        except Exception:
            continue
    if not raw:
        raw = conteudo.decode("utf-8", errors="replace")

    def hex_repl(match):
        value = match.group(1)
        try:
            return bytes.fromhex(value).decode("cp1252", errors="ignore")
        except Exception:
            return ""

    text = re.sub(r"\\'([0-9a-fA-F]{2})", hex_repl, raw)
    text = re.sub(r"\\par[d]?", "\n", text)
    text = re.sub(r"\\tab", "\t", text)
    text = re.sub(r"{\\[^{}]*}", " ", text)
    text = re.sub(r"\\[a-zA-Z]+-?\d* ?", " ", text)
    text = text.replace("{", " ").replace("}", " ").replace("\\", "")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n\s+", "\n", text)
    return text.strip()


def _iterar_arquivos_zip(conteudo: bytes):
    """Percorre ZIPs/pastas compactadas e devolve arquivos treinaveis."""
    try:
        with ZipFile(io.BytesIO(conteudo)) as archive:
            for info in archive.infolist():
                if info.is_dir():
                    continue
                name = info.filename
                lower = name.lower()
                if not lower.endswith((".rtf", ".txt", ".pdf", ".docx", ".xls", ".xlsx")):
                    continue
                yield name, archive.read(info)
    except BadZipFile as exc:
        raise HTTPException(400, "Arquivo ZIP invalido.") from exc


def _criar_documento_motor(db: Session, nome_arquivo: str, conteudo: bytes, categoria: str):
    hash_doc = hashlib.sha256(conteudo).hexdigest()
    existente = db.query(MotorDocumento).filter_by(hash_conteudo=hash_doc).first()
    if existente:
        return {"status": "duplicata", "documento_id": existente.id, "arquivo": nome_arquivo}

    texto = _extrair_texto_arquivo(conteudo, nome_arquivo)
    if len(texto.strip()) < 100:
        return {"status": "ignorado", "arquivo": nome_arquivo, "motivo": "texto extraido muito curto"}

    cat = categoria.upper().strip()
    if cat == "AUTO":
        cat = cerebro_service.detectar_categoria(texto)

    ext = (nome_arquivo.split(".")[-1] if "." in nome_arquivo else "txt").lower()
    doc = MotorDocumento(
        categoria      = cat,
        nome_arquivo   = nome_arquivo,
        tipo_arquivo   = ext,
        texto_extraido = texto[:100_000],
        texto_limpo    = texto[:60_000],
        hash_conteudo  = hash_doc,
        status_analise = "pendente",
    )
    db.add(doc)
    db.flush()
    db.add(MotorFila(documento_id=doc.id, status="aguardando"))
    return {"status": "enfileirado", "documento_id": doc.id, "arquivo": nome_arquivo, "categoria": cat, "chars": len(texto)}


# â”€â”€ Upload de documento â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@router.post("/upload")
async def upload_documento(
    arquivo:   UploadFile = File(...),
    categoria: str = Form("AUTO"),
    db:        Session = Depends(get_db),
):
    """
    Recebe um documento, extrai texto e enfileira para anÃ¡lise.
    Se categoria="AUTO", detecta automaticamente pelo conteÃºdo.
    """
    conteudo = await arquivo.read()
    if not conteudo:
        raise HTTPException(400, "Arquivo vazio")

    filename = arquivo.filename or "arquivo"
    lower = filename.lower()

    if lower.endswith(".zip"):
        resultados = []
        for inner_name, inner_bytes in _iterar_arquivos_zip(conteudo):
            resultados.append(_criar_documento_motor(db, inner_name, inner_bytes, categoria))
        db.commit()
        enfileirados = [item for item in resultados if item.get("status") == "enfileirado"]
        duplicatas = [item for item in resultados if item.get("status") == "duplicata"]
        ignorados = [item for item in resultados if item.get("status") == "ignorado"]
        if not resultados:
            raise HTTPException(422, "ZIP sem arquivos RTF, TXT, PDF ou DOCX treinaveis.")
        return {
            "status": "enfileirado_zip",
            "documentos": len(enfileirados),
            "duplicatas": len(duplicatas),
            "ignorados": len(ignorados),
            "itens": resultados[:50],
            "mensagem": f"ZIP processado: {len(enfileirados)} documento(s) enfileirado(s), {len(duplicatas)} duplicata(s), {len(ignorados)} ignorado(s).",
        }

    resultado = _criar_documento_motor(db, filename, conteudo, categoria)
    db.commit()
    if resultado.get("status") == "ignorado":
        raise HTTPException(422, resultado.get("motivo", "Texto extraido muito curto."))
    if resultado.get("status") == "duplicata":
        return {
            **resultado,
            "mensagem": "Este documento ja foi enviado anteriormente.",
        }
    return {
        **resultado,
        "mensagem": f"Documento classificado como '{resultado.get('categoria')}' e enfileirado para analise.",
    }


@router.post("/upload-com-resultado")
async def upload_com_resultado(
    arquivo:   UploadFile = File(...),
    categoria: str = Form("AUTO"),
    db:        Session = Depends(get_db),
):
    """
    Recebe um documento que inclui resultado final (decisÃ£o/sentenÃ§a).
    Extrai inteligÃªncia estratÃ©gica imediatamente (sem fila).
    Ideal para processos completos, casos com desfecho conhecido.
    """
    conteudo = await arquivo.read()
    if not conteudo:
        raise HTTPException(400, "Arquivo vazio")

    hash_doc = hashlib.sha256(conteudo).hexdigest()
    existente = db.query(MotorAprendizadoEstrategico).filter_by(hash_conteudo=hash_doc).first()
    if existente:
        return {"status": "duplicata", "mensagem": "Este documento jÃ¡ foi analisado anteriormente."}

    texto = _extrair_texto_arquivo(conteudo, arquivo.filename or "arquivo")
    if len(texto.strip()) < 200:
        raise HTTPException(422, "Texto muito curto para anÃ¡lise estratÃ©gica.")

    cat = categoria.upper().strip()
    if cat == "AUTO":
        cat = cerebro_service.detectar_categoria(texto)

    estrategia = cerebro_service.analisar_com_resultado(texto, cat)
    ae = cerebro_service.salvar_aprendizado_estrategico(
        db, cat, arquivo.filename or "", hash_doc, estrategia
    )
    db.commit()

    return {
        "status":      "analisado",
        "categoria":   cat,
        "tipo_pedido": ae.tipo_pedido,
        "resultado":   ae.resultado,
        "licao":       ae.licao_estrategica,
        "mensagem":    f"Documento analisado â€” resultado: {ae.resultado}.",
    }


# â”€â”€ Consultas de conhecimento â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@router.get("/status")
def status_geral():
    """Bateria de conhecimento de todas as categorias."""
    return cerebro_service.obter_status_todos()


@router.get("/arvore")
def arvore_conhecimento():
    """Ãrvore completa de conhecimento com dados estratÃ©gicos."""
    return cerebro_service.obter_arvore_conhecimento()


@router.get("/perfil/{categoria}")
def perfil_categoria(categoria: str):
    """PadrÃµes de estilo aprendidos para uma categoria."""
    cat = categoria.upper()
    if cat not in CATEGORIAS_VALIDAS:
        raise HTTPException(400, f"Categoria invÃ¡lida. VÃ¡lidas: {list(CATEGORIAS_VALIDAS)}")
    perfil = cerebro_service.obter_perfil_categoria(cat)
    if not perfil:
        return {"categoria": cat, "percentual": 0, "total_docs": 0,
                "mensagem": "Nenhum documento analisado ainda."}
    return perfil


@router.get("/estrategia/{categoria}")
def estrategia_categoria(categoria: str, tipo_pedido_slug: str = None):
    """InteligÃªncia estratÃ©gica consolidada para uma categoria."""
    cat = categoria.upper()
    if cat not in CATEGORIAS_VALIDAS:
        raise HTTPException(400, f"Categoria invÃ¡lida. VÃ¡lidas: {list(CATEGORIAS_VALIDAS)}")
    return cerebro_service.obter_inteligencia(cat, tipo_pedido_slug)


# â”€â”€ Fila e processamento â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@router.get("/fila")
def status_fila(db: Session = Depends(get_db)):
    """Status da fila de processamento."""
    from sqlalchemy import func
    contagens = dict(
        db.query(MotorFila.status, func.count(MotorFila.id))
        .group_by(MotorFila.status).all()
    )
    itens = (
        db.query(MotorFila, MotorDocumento)
        .join(MotorDocumento, MotorFila.documento_id == MotorDocumento.id)
        .order_by(MotorFila.criado_em.desc())
        .limit(20).all()
    )
    return {
        "aguardando":  contagens.get("aguardando", 0),
        "processando": contagens.get("processando", 0),
        "concluido":   contagens.get("concluido", 0),
        "falha":       contagens.get("falha", 0),
        "itens": [
            {
                "id":         f.id,
                "arquivo":    d.nome_arquivo,
                "categoria":  d.categoria,
                "status":     f.status,
                "tentativas": f.tentativas,
                "erro":       f.erro_msg,
                "criado_em":  f.criado_em.isoformat() if f.criado_em else None,
            }
            for f, d in itens
        ],
    }


@router.post("/processar-agora")
def processar_agora(db: Session = Depends(get_db)):
    """
    ForÃ§a o processamento imediato da fila (sem esperar o scheduler).
    Ãštil para testar ou forÃ§ar reprocessamento urgente.
    """
    pendentes = (
        db.query(MotorFila)
        .filter(MotorFila.status.in_(["aguardando", "falha"]))
        .limit(5).all()
    )
    if not pendentes:
        return {"mensagem": "Nenhum item na fila.", "processados": 0}

    resultados = []
    for item in pendentes:
        item.status     = "processando"
        item.tentativas += 1
        db.commit()
        resultado = cerebro_service.processar_documento(item.documento_id)
        item.status   = "concluido" if resultado.get("ok") else "falha"
        item.erro_msg = resultado.get("erro", "")
        db.commit()
        resultados.append({"documento_id": item.documento_id, **resultado})

    return {"processados": len(resultados), "resultados": resultados}


# â”€â”€ Grafo de Conhecimento â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@router.get("/grafo/status")
def grafo_status():
    """EstatÃ­sticas do Grafo de Conhecimento."""
    return kg_service.status_grafo()


@router.get("/grafo/{categoria}")
def grafo_categoria(categoria: str, tipo_pedido: str = ""):
    """InteligÃªncia consolidada do grafo para uma categoria."""
    cat = categoria.upper()
    if cat not in CATEGORIAS_VALIDAS:
        raise HTTPException(400, f"Categoria invÃ¡lida. VÃ¡lidas: {list(CATEGORIAS_VALIDAS)}")
    return kg_service.consultar_grafo_para_geracao(cat, tipo_pedido)


@router.post("/grafo/recalcular-pesos")
def recalcular_pesos():
    """Recalcula pesos temporais (execute semanalmente via cron)."""
    return kg_service.recalcular_pesos_temporais()


# â”€â”€ Feedback â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@router.post("/feedback/{doc_id}")
def registrar_feedback(doc_id: int, body: dict):
    """
    Registra feedback do usuÃ¡rio sobre um documento gerado.

    Body:
        feedback:       "usou" | "editou" | "descartou"
        nota:           1-5 (opcional)
        resultado_real: "favoravel" | "parcial" | "desfavoravel" (opcional)
        obs:            texto livre (opcional)
    """
    resultado = kg_service.processar_feedback(
        minuta_id=doc_id,
        feedback=body.get("feedback", "usou"),
        nota=body.get("nota"),
        resultado_real=body.get("resultado_real"),
        obs=body.get("obs"),
    )
    if not resultado.get("ok"):
        raise HTTPException(400, resultado.get("erro", "Erro ao registrar feedback"))
    return resultado


@router.get("/minutas")
def listar_minutas(
    categoria: str = None,
    feedback:  str = None,
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
):
    """Lista documentos gerados com mÃ©tricas de uso e feedback."""
    q = db.query(MotorMinuta)
    if categoria:
        q = q.filter_by(categoria=categoria)
    if feedback:
        q = q.filter_by(feedback=feedback)
    total = q.count()
    itens = q.order_by(MotorMinuta.criado_em.desc()).offset(skip).limit(limit).all()
    return {
        "total": total,
        "itens": [
            {
                "id":            m.id,
                "categoria":     m.categoria,
                "modelo_label":  m.modelo_label,
                "feedback":      m.feedback,
                "nota":          m.feedback_nota,
                "resultado_real": m.resultado_real,
                "criado_em":     m.criado_em.isoformat() if m.criado_em else None,
            }
            for m in itens
        ],
    }

